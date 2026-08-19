import openpyxl

files = {
    'AI_Study_Planner_Selenium_E2E.xlsx': ['Web E2E Dashboard', 'All Selenium E2E Test Cases'],
    'AI_Study_Planner_UI_UX_Test.xlsx': ['UI UX Dashboard', 'All UI UX Test Cases'],
    'AI_Study_Planner_Load_Test.xlsx': ['Load Test Dashboard', 'All 300+ Load Test Cases'],
    'AI_Study_Planner_Appium_E2E.xlsx': ['Mobile E2E Dashboard', 'Appium Mobile Test Cases'],
    'AI_Study_Planner_MASTER_Test_Report.xlsx': ['Executive Summary', 'UI UX Testing', 'Selenium Web E2E', 'Load Testing', 'Appium Mobile E2E', 'Master Test Case Index', 'Test Statistics', 'Production Readiness']
}

for fname, exp_sheets in files.items():
    path = f'testing/reports/{fname}'
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f'Checking {fname}:')
    for s in exp_sheets:
        if s in wb.sheetnames:
            ws = wb[s]
            print(f'  [OK] Sheet "{s}": {ws.max_row} rows')
        else:
            print(f'  [FAIL] Sheet "{s}" MISSING!')
